"""Generate the Excel workbook with financial statements, metrics, charts, and metadata."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter


# Styles
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

LABEL_FONT = Font(name="Calibri", bold=True, size=10)
VALUE_FONT = Font(name="Calibri", size=10)
PCT_FONT = Font(name="Calibri", size=10, color="333333")

THIN_BORDER = Border(
    bottom=Side(style="thin", color="D0D0D0"),
)

SECTION_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

USD_FORMAT = '#,##0;(#,##0);"-"'
USD_MILLIONS_FORMAT = '#,##0;(#,##0);"-"'
SHARES_FORMAT = '#,##0;(#,##0);"-"'
PCT_FORMAT = '0.0%;(0.0%);"-"'
EPS_FORMAT = '#,##0.00;(#,##0.00);"-"'
RATIO_FORMAT = '0.00x;(0.00x);"-"'


def _apply_header_style(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border = THIN_BORDER


def _get_format(line_item: str, sheet_name: str) -> str:
    lower = line_item.lower()
    if "eps" in lower:
        return EPS_FORMAT
    if "share" in lower:
        return SHARES_FORMAT
    if "margin" in lower or "growth" in lower or "cagr" in lower or "change" in lower:
        return PCT_FORMAT
    if "debt / equity" in lower:
        return RATIO_FORMAT
    if "quality" in lower:
        return "@"  # text
    return USD_FORMAT


def _write_statement_sheet(
    wb: Workbook,
    sheet_name: str,
    period_dates: list[str],
    data: dict[str, list],
    ticker: str,
):
    """Write a single financial statement sheet."""
    ws = wb.create_sheet(title=sheet_name)
    n_periods = len(period_dates)

    # Title row
    ws.cell(row=1, column=1, value=f"{ticker} — {sheet_name.replace('_', ' ')}")
    ws.cell(row=1, column=1).font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_periods + 1)

    # Header row with period dates
    header_row = 3
    ws.cell(row=header_row, column=1, value="Line Item")
    _apply_header_style(ws.cell(row=header_row, column=1))

    for j, date in enumerate(period_dates):
        cell = ws.cell(row=header_row, column=j + 2, value=date)
        _apply_header_style(cell)

    # Data rows
    row = header_row + 1
    for line_item, values in data.items():
        label_cell = ws.cell(row=row, column=1, value=line_item)
        label_cell.font = LABEL_FONT
        label_cell.border = THIN_BORDER

        fmt = _get_format(line_item, sheet_name)

        for j in range(n_periods):
            val = values[j] if j < len(values) else None
            cell = ws.cell(row=row, column=j + 2)
            if val is not None:
                cell.value = val
                cell.number_format = fmt
            else:
                cell.value = "—"
                cell.alignment = Alignment(horizontal="center")
            cell.font = VALUE_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="right")

        row += 1

    # Auto-width columns
    ws.column_dimensions["A"].width = 32
    for j in range(n_periods):
        ws.column_dimensions[get_column_letter(j + 2)].width = 18

    # Freeze panes
    ws.freeze_panes = "B4"


def _write_metrics_sheet(
    wb: Workbook,
    period_dates: list[str],
    metrics: dict[str, list],
    ticker: str,
):
    """Write the Key_Ratios sheet."""
    ws = wb.create_sheet(title="Key_Ratios")
    n_periods = len(period_dates)

    ws.cell(row=1, column=1, value=f"{ticker} — Key Ratios & Metrics")
    ws.cell(row=1, column=1).font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_periods + 1)

    header_row = 3
    ws.cell(row=header_row, column=1, value="Metric")
    _apply_header_style(ws.cell(row=header_row, column=1))
    for j, date in enumerate(period_dates):
        cell = ws.cell(row=header_row, column=j + 2, value=date)
        _apply_header_style(cell)

    row = header_row + 1
    for metric_name, values in metrics.items():
        label_cell = ws.cell(row=row, column=1, value=metric_name)
        label_cell.font = LABEL_FONT
        label_cell.border = THIN_BORDER

        fmt = _get_format(metric_name, "Key_Ratios")

        for j in range(n_periods):
            val = values[j] if j < len(values) else None
            cell = ws.cell(row=row, column=j + 2)
            if val is not None:
                cell.value = val
                cell.number_format = fmt
            else:
                cell.value = "—"
                cell.alignment = Alignment(horizontal="center")
            cell.font = VALUE_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="right")

        row += 1

    ws.column_dimensions["A"].width = 28
    for j in range(n_periods):
        ws.column_dimensions[get_column_letter(j + 2)].width = 18
    ws.freeze_panes = "B4"


def _write_charts_sheet(
    wb: Workbook,
    charts: list[dict],
    ticker: str,
):
    """Write the Charts sheet with embedded PNGs."""
    ws = wb.create_sheet(title="Charts")

    ws.cell(row=1, column=1, value=f"{ticker} — Charts")
    ws.cell(row=1, column=1).font = Font(name="Calibri", bold=True, size=13, color="1F4E79")

    row = 3
    for chart_info in charts:
        path = chart_info["path"]
        title = chart_info["title"]

        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=11)
        row += 1

        if Path(path).exists():
            img = XlImage(str(path))
            img.width = 720
            img.height = 360
            ws.add_image(img, f"A{row}")
            row += 22  # Leave space for the image
        else:
            ws.cell(row=row, column=1, value=f"(Chart image not found: {path})")
            row += 2


def _write_metadata_sheet(
    wb: Workbook,
    run_config: dict,
    warnings: list[str],
    data_urls: list[str],
    ticker: str,
):
    """Write the Run_Metadata sheet."""
    ws = wb.create_sheet(title="Run_Metadata")

    ws.cell(row=1, column=1, value=f"{ticker} — Run Metadata")
    ws.cell(row=1, column=1).font = Font(name="Calibri", bold=True, size=13, color="1F4E79")

    row = 3
    meta_items = [
        ("Ticker", run_config.get("ticker", "")),
        ("Period Type", run_config.get("period_type", "")),
        ("Number of Periods", run_config.get("num_periods", "")),
        ("Mapping Version", run_config.get("mapping_version", "")),
        ("Generated At", datetime.utcnow().isoformat()),
    ]

    for label, value in meta_items:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=2, value=str(value)).font = VALUE_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Data Sources").font = Font(bold=True, size=11)
    row += 1
    for url in data_urls:
        ws.cell(row=row, column=1, value=url).font = VALUE_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Warnings").font = Font(bold=True, size=11)
    row += 1
    if warnings:
        for w in warnings:
            ws.cell(row=row, column=1, value=w).font = Font(size=9, color="CC6600")
            row += 1
    else:
        ws.cell(row=row, column=1, value="None").font = VALUE_FONT

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60


def export_workbook(
    ticker: str,
    period_type: str,
    period_dates: list[str],
    normalized: dict[str, dict[str, list]],
    metrics: dict[str, list],
    charts: list[dict],
    run_config: dict,
    warnings: list[str],
    data_urls: list[str],
    output_dir: Path,
) -> Path:
    """Generate the complete Excel workbook and return the file path."""
    wb = Workbook()

    # Remove the default sheet
    wb.remove(wb.active)

    # Write statement sheets
    for sheet_name in ["Income_Statement", "Balance_Sheet", "Cash_Flow"]:
        if sheet_name in normalized:
            _write_statement_sheet(wb, sheet_name, period_dates, normalized[sheet_name], ticker)

    # Key ratios
    _write_metrics_sheet(wb, period_dates, metrics, ticker)

    # Charts
    if charts:
        _write_charts_sheet(wb, charts, ticker)

    # Metadata
    _write_metadata_sheet(wb, run_config, warnings, data_urls, ticker)

    # Save
    date_str = datetime.utcnow().strftime("%Y%m%d")
    filename = f"{ticker}_{period_type.replace('-', '')}_{date_str}.xlsx"
    output_path = output_dir / filename
    output_dir.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    return output_path
